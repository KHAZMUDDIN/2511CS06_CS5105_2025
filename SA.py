import pandas as pd
from collections import defaultdict

from pprint import pprint
import glob

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from math import ceil
from datetime import datetime

import pandas as pd
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
import logging


# -------------------------------------------------
# Helper: from students in a room -> guess course code
# -------------------------------------------------
def infer_course_code_for_room(result, date_str, shift, room_df):
    """
    result[date_str][shift][course]["df"] has full list for each course.
    We pick the course with maximum overlap of roll numbers.
    """
    slot_courses = result[date_str][shift]
    room_rolls = set(room_df["rollno"])

    best_course = None
    best_overlap = 0

    for code, info in slot_courses.items():
        course_rolls = set(info["df"]["rollno"])
        overlap = len(room_rolls & course_rolls)
        if overlap > best_overlap:
            best_overlap = overlap
            best_course = code

    return best_course


# -------------------------------------------------
# Helper: write ONE excel file for ONE room
# -------------------------------------------------
def write_room_excel(
    base_output_dir,
    date_str,
    day,
    shift,
    room_no,
    block,
    course_code,
    students_df,
):
    """
    Creates folder structure and writes one Excel file:
    output/date/shift/course_code/dd_mm_yyyy_course_room_shift.xlsx
    """
    # folder names
    date_folder_name = date_str.replace("-", "_")           # dd-mm-yyyy -> dd_mm_yyyy
    date_dir = os.path.join(base_output_dir, date_folder_name)
    shift_dir = os.path.join(date_dir, shift)
    course_dir = os.path.join(shift_dir, course_code)

    os.makedirs(course_dir, exist_ok=True)

    # file name
    filename = f"{date_folder_name}_{course_code}_{room_no}_{shift}.xlsx"
    filepath = os.path.join(course_dir, filename)

    # ---- create workbook ----
    wb = Workbook()
    ws = wb.active
    ws.title = f"{course_code} Room {room_no}"

    # merged header cell (row 1, cols A-C)
    header_text = (
        f"Course: {course_code} | Room: {room_no} | Date: {date_str} | Session: {shift}"
    )
    ws.merge_cells("A1:C1")
    ws["A1"] = header_text
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # column headers (row 2)
    ws["A2"] = "Roll"
    ws["B2"] = "Student Name"
    ws["C2"] = "Signature"
    ws["A2"].font = ws["B2"].font = ws["C2"].font = Font(bold=True)

    # write students starting from row 3
    row_start = 3
    for i, (_, row) in enumerate(students_df.iterrows(), start=row_start):
        ws.cell(row=i, column=1, value=row["rollno"])
        ws.cell(row=i, column=2, value=row["Name"])
        # Signature column left blank (column 3)

    # TA rows (5 rows)
    ta_start = row_start + len(students_df)
    for idx in range(5):
        ws.cell(row=ta_start + idx, column=1, value=f"TA{idx+1}")

    # Invigilator rows (5 rows)
    inv_start = ta_start + 5
    for idx in range(5):
        ws.cell(row=inv_start + idx, column=1, value=f"Invigilator{idx+1}")

    # adjust column widths a bit
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18

    wb.save(filepath)
    # print(f"Saved: {filepath}")


# -------------------------------------------------
# Main: generate all room-wise Excel files for ONE date+shift
# room_alloc: dict returned by allocate_shift_for_date(...)
# -------------------------------------------------
def export_shift_excels(result, date_str, shift, room_alloc, base_output_dir="output"):
    """
    result      : big dict with all course data
    date_str    : 'dd-mm-yyyy'
    shift       : 'Morning' or 'Evening'
    room_alloc  : dict from allocate_shift_for_date()
                  room_no -> {date, day, shift, room_no, block, count, df}
    base_output_dir : root 'output' folder
    """
    if not room_alloc:
        return

    # day is same for all rooms on this date/shift
    any_room = next(iter(room_alloc.values()))
    day_val = any_room["day"]

    for room_no, info in room_alloc.items():
        students_df = info["df"]

        # find course_code for this room using rolls
        course_code = infer_course_code_for_room(result, date_str, shift, students_df)
        if course_code is None:
            # fallback if something weird: put under 'MIXED'
            course_code = "MIXED"

        write_room_excel(
            base_output_dir=base_output_dir,
            date_str=date_str,
            day=day_val,
            shift=shift,
            room_no=room_no,
            block=info["block"],
            course_code=course_code,
            students_df=students_df,
        )




def add_names(df_rolls, roll_name):
    # 2️⃣ Clean both roll columns
    df_rolls["rollno"] = df_rolls["rollno"].astype(str).str.strip()
    roll_name["Roll"] = roll_name["Roll"].astype(str).str.strip()
    roll_name["Name"] = roll_name["Name"].astype(str).str.strip()

    # 3️⃣ Merge to add the name column
    df_with_names = df_rolls.merge(
        roll_name,
        left_on="rollno",
        right_on="Roll",
        how="left"
    )

    # 4️⃣ Keep only rollno + name, fill missing names
    df_with_names = df_with_names[["rollno", "Name"]]
    df_with_names["Name"] = df_with_names["Name"].fillna("Unknown Name")

    # print(df_with_names)
    return df_with_names


def get_rollnos_for_course(course_roll, course_code):
    # normalize for safety
    course_roll["course_code"] = course_roll["course_code"].astype(str).str.strip()
    course_roll["rollno"] = course_roll["rollno"].astype(str).str.strip()

    # filter rows for this course
    filtered = course_roll[course_roll["course_code"] == course_code]

    # create dataframe with only rollno column
    df = filtered[["rollno"]].reset_index(drop=True)

    return df

def build_course_dicts(excel_path):
    # Load sheets
    tt = pd.read_excel(excel_path, sheet_name="in_timetable")
    course_roll = pd.read_excel(excel_path, sheet_name="in_course_roll_mapping")
    roll_name = pd.read_excel(excel_path, sheet_name="in_roll_name_mapping")

    # ---- adjust these column names to match your file exactly ----
    date_col = "Date"                # column in in_timetable with the exam date
    day_col = "Day"
    morning_col = "Morning"          # column with morning course codes
    evening_col = "Evening"          # column with evening course codes

    course_col = "course_code"        # column in in_course_roll_mapping with course code
    roll_col = "rollno"          # column in in_course_roll_mapping with roll number

    roll_name_roll_col = "Roll"  # column in in_roll_name_mapping with roll number
    name_col = "Name"                  # column in in_roll_name_mapping with student name
    # ---------------------------------------------------------------

    # Clean up whitespace in important columns
    course_roll[course_col] = course_roll[course_col].astype(str).str.strip()
    course_roll[roll_col] = course_roll[roll_col].astype(str).str.strip()
    roll_name[roll_name_roll_col] = roll_name[roll_name_roll_col].astype(str).str.strip()

    # tt sheet[morning_col] first row element
    # print(tt[morning_col][0])
    # print(type(tt[morning_col][0]))

    #==================================================================
    # # s = "CS249; CH426; MM304; CB308; CE216; CB204; PH422; MM202"
    #
    # s = tt[morning_col][0]
    #
    # course_list = [c.strip() for c in s.split(";") if c.strip()]
    #
    # # all courses of first date 4/30/2016 morning shift in a list
    # print(course_list)
    # # first course
    # # print(course_list[0])
    #
    # course_code = course_list[0]
    # print(course_code)
    # rolls_of_course_code = get_rollnos_for_course(course_roll, course_code)
    #
    # roll_name_of_course_code = add_names(rolls_of_course_code, roll_name)
    #
    # # print(roll_name_of_course_code)
    #
    # df_sorted = roll_name_of_course_code.sort_values(by="rollno").reset_index(drop=True)
    # print(df_sorted)
    #====================================================================

    # # columns in tt for date and day – change if needed
    # date_col = "Date"
    # day_col = "Day"  # or "Term" / "DayName" etc, whatever your sheet uses
    #
    # row_idx = 0  # row of tt you are working with (here: first date)
    #
    # # all courses of that date, morning shift
    # s = tt[morning_col][row_idx]
    # course_list = [c.strip() for c in str(s).split(";") if c.strip()]
    #
    # date_val = tt.loc[row_idx, date_col]
    # day_val = tt.loc[row_idx, day_col]
    #
    # for course_code in course_list:
    #     print(f"\n==============================")
    #     print(f"Date: {date_val}  Day: {day_val}")
    #     print(f"Shift: Morning  Course: {course_code}")
    #     print(f"==============================")
    #
    #     rolls_of_course_code = get_rollnos_for_course(course_roll, course_code)
    #     roll_name_of_course_code = add_names(rolls_of_course_code, roll_name)
    #     df_sorted = roll_name_of_course_code.sort_values(by="rollno").reset_index(drop=True)
    #
    #     print(df_sorted)

    #==================================================================
    # # column names — change if your sheet uses different ones
    # date_col = "Date"
    # day_col = "Day"
    # morning_col = "Morning"
    # evening_col = "Evening"
    #
    #
    # for idx, row in tt.iterrows():
    #
    #     date_val = row[date_col]
    #     day_val = row[day_col]
    #
    #     print("\n====================================================")
    #     print(f"DATE: {date_val}   DAY: {day_val}")
    #     print("====================================================\n")
    #
    #     # ----------- MORNING SHIFT ------------
    #     morning_entry = row[morning_col]
    #     if pd.notna(morning_entry):
    #         morning_courses = [c.strip() for c in str(morning_entry).split(";") if c.strip()]
    #
    #         print("\n********** MORNING SHIFT **********\n")
    #         for course_code in morning_courses:
    #             print(f"\n--- Course: {course_code} ---")
    #
    #             rolls_df = get_rollnos_for_course(course_roll, course_code)
    #             names_df = add_names(rolls_df, roll_name)
    #             df_sorted = names_df.sort_values(by="rollno").reset_index(drop=True)
    #
    #             # print(df_sorted)
    #
    #     # ----------- EVENING SHIFT ------------
    #     evening_entry = row[evening_col]
    #     if pd.notna(evening_entry):
    #         evening_courses = [c.strip() for c in str(evening_entry).split(";") if c.strip()]
    #
    #         print("\n********** EVENING SHIFT **********\n")
    #         for course_code in evening_courses:
    #             print(f"\n--- Course: {course_code} ---")
    #
    #             rolls_df = get_rollnos_for_course(course_roll, course_code)
    #             names_df = add_names(rolls_df, roll_name)
    #             df_sorted = names_df.sort_values(by="rollno").reset_index(drop=True)
    #
    #             # print(df_sorted)
    #
    #     # break
    all_courses_data = {}

    date_col = "Date"
    day_col = "Day"
    morning_col = "Morning"
    evening_col = "Evening"

    for idx, row in tt.iterrows():

        # convert date to string dd-mm-yyyy
        date_val = pd.to_datetime(row[date_col]).strftime("%d-%m-%Y")
        day_val = row[day_col]

        # create the dictionary for this date
        all_courses_data[date_val] = {
            "Morning": {},
            "Evening": {}
        }

        # -------- MORNING SHIFT --------
        morning_entry = row[morning_col]
        if pd.notna(morning_entry):
            morning_courses = [c.strip() for c in str(morning_entry).split(";") if c.strip()]

            for course_code in morning_courses:
                rolls_df = get_rollnos_for_course(course_roll, course_code)
                names_df = add_names(rolls_df, roll_name)
                df_sorted = names_df.sort_values(by="rollno").reset_index(drop=True)

                all_courses_data[date_val]["Morning"][course_code] = {
                    "date": date_val,
                    "day": day_val,
                    "shift": "Morning",
                    "course_code": course_code,
                    "df": df_sorted,
                    "count": len(df_sorted)
                }

        # -------- EVENING SHIFT --------
        evening_entry = row[evening_col]
        if pd.notna(evening_entry):
            evening_courses = [c.strip() for c in str(evening_entry).split(";") if c.strip()]

            for course_code in evening_courses:
                rolls_df = get_rollnos_for_course(course_roll, course_code)
                names_df = add_names(rolls_df, roll_name)
                df_sorted = names_df.sort_values(by="rollno").reset_index(drop=True)

                all_courses_data[date_val]["Evening"][course_code] = {
                    "date": date_val,
                    "day": day_val,
                    "shift": "Evening",
                    "course_code": course_code,
                    "df": df_sorted,
                    "count": len(df_sorted)
                }

    return all_courses_data
#================================================================


#Pretty-print just the structure (not the full dataframes)
def preview_structure1(data):
    clean_view = {}

    for date, shifts in data.items():
        clean_view[date] = {
            shift: list(courses.keys())
            for shift, courses in shifts.items()
        }

    pprint(clean_view)

# Print only counts per course (no dataframe)
def preview_counts(data):
    for date, shifts in data.items():
        print(f"\n===== {date} =====")
        for shift, courses in shifts.items():
            print(f"\n--- {shift} ---")
            for course, info in courses.items():
                print(f"{course}: {info['count']} students")

# Print summary for a single date (clean)
def preview_date(data, date):
    if date not in data:
        print("Date not found.")
        return

    print(f"\n===== DATE: {date} =====")
    for shift, courses in data[date].items():
        print(f"\n--- {shift} ---")
        for course, info in courses.items():
            print(f"{course}: {info['count']} students")
# Print only course names + student counts in table-like format
def preview_table(data):
    for date, shifts in data.items():
        print(f"\n📅 {date}")
        print("-" * 40)
        for shift, courses in shifts.items():
            print(f"\n  {shift}:")
            for course, info in courses.items():
                print(f"    {course:<10}  -> {info['count']} students")

def detect_clashes(result):
    clashes = []   # will store: (date, shift, courseA, courseB, [common rolls])

    for date, shifts in result.items():

        for shift, courses in shifts.items():

            course_codes = list(courses.keys())

            # compare each pair of courses in the same shift
            for i in range(len(course_codes)):
                for j in range(i + 1, len(course_codes)):

                    c1 = course_codes[i]
                    c2 = course_codes[j]

                    rolls1 = set(courses[c1]["df"]["rollno"])
                    rolls2 = set(courses[c2]["df"]["rollno"])

                    common = rolls1.intersection(rolls2)

                    if common:   # clash found
                        clashes.append({
                            "date": date,
                            "shift": shift,
                            "course1": c1,
                            "course2": c2,
                            "common_rolls": sorted(list(common)),
                            "count": len(common)
                        })

    return clashes

# result  -> your big dictionary (date -> shift -> course -> {...})
# rooms_df -> dataframe with room capacities (we'll use this later)

def get_sorted_courses_for_slot(result, date_str, shift):
    """
    For a given date (e.g. '30-04-2016') and shift ('Morning' or 'Evening'),
    return a list of dicts sorted by student count (largest first).

    Each element:
    {
        "course_code": ...,
        "count": ...,
        "df": ...   # dataframe of rollno + Name
    }
    """
    slot_data = result[date_str][shift]   # dict: course_code -> info

    courses = []
    for course_code, info in slot_data.items():
        courses.append({
            "course_code": course_code,
            "count": info["count"],
            "df": info["df"]
        })

    # sort by count descending
    courses_sorted = sorted(courses, key=lambda x: x["count"], reverse=True)
    return courses_sorted


# example usage (for one date and Morning slot):
# sorted_courses = get_sorted_courses_for_slot(result, "30-04-2016", "Morning")
# for c in sorted_courses:
#     print(c["course_code"], c["count"])

def prepare_rooms(rooms_df, buffer):
    df = rooms_df.copy()

    # effective capacity after buffer
    df["effective_capacity"] = (df["Exam Capacity"] - buffer).clip(lower=0)

    # sort within each Block by capacity (high to low)
    df_sorted = df.sort_values(
        by=["Block", "Exam Capacity"],
        ascending=[True, False]
    ).reset_index(drop=True)

    return df_sorted

# ---- check if total capacity is enough for a given date + shift ----
def can_allocate_shift(result, date_str, shift, rooms_sorted):
    # total students in this shift
    total_students = sum(
        info["count"] for info in result[date_str][shift].values()
    )

    # total effective capacity across all rooms
    total_capacity = rooms_sorted["effective_capacity"].sum()

    print(f"Date: {date_str}, Shift: {shift}")
    print(f"Total students      : {total_students}")
    print(f"Total eff. capacity : {total_capacity}")

    if total_capacity >= total_students:
        print("Allocation is POSSIBLE for this shift.")
        return True
    else:
        print("Allocation is NOT POSSIBLE for this shift.")
        return False
def effective_capacity(capacity, buffer, mode):
    """Return effective capacity of a room for ONE subject."""
    eff = max(0, capacity - buffer)
    mode = mode.lower()
    if mode == "sparse":
        return eff // 2
    elif mode == "dense":
        return eff
    else:
        raise ValueError("mode must be 'Sparse' or 'Dense'")
def allocate_shift_for_date(result, date_str, shift, rooms_df, buffer, mode):
    """
    Allocate all courses for a given date + shift into rooms.

    result   : big dict (date -> shift -> course -> {...})
    date_str : 'dd-mm-yyyy' (e.g. '30-04-2016')
    shift    : 'Morning' or 'Evening'
    rooms_df : dataframe of in_room_capacity
    buffer   : int
    mode     : 'Sparse' or 'Dense'

    Returns:
        room_alloc = {
            room_no: {
                "date": ...,
                "day": ...,
                "shift": ...,
                "room_no": ...,
                "block": ...,
                "count": total_students_in_room,
                "df": dataframe(rollno, Name)
            },
            ...
        }
    """
    # ---------- 0. basic data ----------
    slot_data = result[date_str][shift]     # course_code -> info
    if not slot_data:
        return {}

    # same for all courses on that date
    sample_info = next(iter(slot_data.values()))
    day_val = sample_info["day"]

    # courses sorted by size (largest first)
    courses_sorted = sorted(
        slot_data.items(),
        key=lambda kv: kv[1]["count"],
        reverse=True
    )
    # print(courses_sorted)
 # ---------- 1. prepare rooms ----------
    df = rooms_df.copy()
    df["Room No."] = df["Room No."].astype(str).str.strip()
    df["Block"] = df["Block"].astype(str).str.strip()
    df["Exam Capacity"] = df["Exam Capacity"].astype(int)

    df["effective_capacity"] = df["Exam Capacity"].apply(
        lambda c: effective_capacity(c, buffer, mode)
    )
    df = df[df["effective_capacity"] > 0]

    # sort by Block then Room No. for adjacency
    df = df.sort_values(["Block", "Room No."]).reset_index(drop=True)

    # room state
    rooms = []
    for _, r in df.iterrows():
        rooms.append({
            "room_no": r["Room No."],
            "block": r["Block"],
            "capacity": int(r["Exam Capacity"]),
            "eff_cap": int(r["effective_capacity"]),
            "remaining": int(r["effective_capacity"])
        })

    # group rooms by block
    rooms_by_block = defaultdict(list)
    for r in rooms:
        rooms_by_block[r["block"]].append(r)

    # print(rooms_by_block)

    # ---------- 2. allocation helpers ----------

    def can_block_fit_course(block_rooms, need):
        return sum(r["remaining"] for r in block_rooms) >= need

    def allocate_course(course_code, course_info):
        nonlocal rooms, rooms_by_block

        need = course_info["count"]
        if need == 0:
            return

        # choose best block (can fit entire course with min rooms)
        best_block = None
        best_room_count = None

        for block, block_rooms in rooms_by_block.items():
            if not can_block_fit_course(block_rooms, need):
                continue

            # simulate greedy fill in this block to count rooms needed
            remaining = need
            used_rooms = 0
            for r in block_rooms:
                if remaining <= 0:
                    break
                take = min(remaining, r["remaining"])
                if take > 0:
                    used_rooms += 1
                    remaining -= take

            if remaining <= 0:
                if best_block is None or used_rooms < best_room_count:
                    best_block = block
                    best_room_count = used_rooms

        # pick candidate rooms
        if best_block is not None:
            candidate_rooms = rooms_by_block[best_block]
        else:
            # must use multiple blocks
            candidate_rooms = rooms

        # student list for this course
        student_df = course_info["df"].reset_index(drop=True)
        idx = 0  # pointer into course dataframe

        # assign students to rooms
        for r in candidate_rooms:
            if need <= 0:
                break
            if r["remaining"] <= 0:
                continue

            take = min(need, r["remaining"])
            if take <= 0:
                continue

            # slice students for this room
            assigned = student_df.iloc[idx: idx + take].copy()
            idx += take
            need -= take
            r["remaining"] -= take

            # store in room_alloc dict
            key = r["room_no"]
            if key not in room_alloc:
                room_alloc[key] = {
                    "date": date_str,
                    "day": day_val,
                    "shift": shift,
                    "room_no": r["room_no"],
                    "block": r["block"],
                    "count": 0,
                    "dfs": []   # temp list, will concat later
                }

            room_alloc[key]["dfs"].append(assigned)
            room_alloc[key]["count"] += take

        if need > 0:
            raise ValueError(
                f"Not enough remaining capacity to allocate course {course_code} "
                f"on {date_str} {shift}. Still need {need} students."
            )

    # ---------- 3. main allocation loop ----------
    room_alloc = {}

    for course_code, info in courses_sorted:
        allocate_course(course_code, info)

    # ---------- 4. finalize dfs per room ----------
    for room_no, info in room_alloc.items():
        info["df"] = pd.concat(info.pop("dfs"), ignore_index=True)

    return room_alloc


def get_photo_path(roll, photos_dir):
    """Try multiple extensions, fall back to nopic.jpg."""
    roll = str(roll).strip()
    patterns = [
        os.path.join(photos_dir, f"{roll}.jpg"),
        os.path.join(photos_dir, f"{roll}.jpeg"),
        os.path.join(photos_dir, f"{roll}.png"),
        os.path.join(photos_dir, f"{roll}.*"),   # any extension
    ]

    for pattern in patterns:
        matches = glob.glob(pattern)
        if matches:
            return matches[0]

    # fallback
    return os.path.join(photos_dir, "nopic.jpg")

def generate_overall_seating_excel(result, rooms_df, buffer, mode,
                                   output_dir="output",
                                   filename="op_overall_seating_arrangement.xlsx"):
    """
    Creates a single Excel file in `output_dir` with columns:
    Date, Day, course_code, Room, Allocated_students_count, Roll_list
    """
    rows = []

    for date_str in result.keys():
        for shift in ["Morning", "Evening"]:
            # allocate rooms for this date + shift
            room_alloc = allocate_shift_for_date(
                result, date_str, shift, rooms_df, buffer, mode
            )
            if not room_alloc:
                continue

            # day is same for all rooms in this date+shift
            any_room = next(iter(room_alloc.values()))
            day_val = any_room["day"]

            for room_no, info in room_alloc.items():
                df_room = info["df"].reset_index(drop=True)

                # get course code for this room
                course_code = infer_course_code_for_room(
                    result, date_str, shift, df_room
                )
                if course_code is None:
                    course_code = "MIXED"

                roll_list = ";".join(df_room["rollno"].astype(str))
                count = len(df_room)

                rows.append({
                    "Date": date_str,
                    "Day": day_val,
                    "course_code": course_code,
                    "Room": room_no,
                    "Allocated_students_count": count,
                    "Roll_list": roll_list
                })

    if not rows:
        print("No allocations found; not creating overall seating file.")
        return

    df_out = pd.DataFrame(
        rows,
        columns=[
            "Date",
            "Day",
            "course_code",
            "Room",
            "Allocated_students_count",
            "Roll_list",
        ],
    )

    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, filename)
    df_out.to_excel(out_path, index=False, sheet_name="Seating Plan")
    print(f"Overall seating file written to: {out_path}")

#======================================
# ---------- write ONE excel file for ONE (course, room) ----------
def write_room_excel(
    base_output_dir,
    date_str,
    day,
    shift,
    room_no,
    block,
    course_code,
    students_df,
):
    # folders: output / dd_mm_yyyy / shift /
    date_folder_name = date_str.replace("-", "_")        # dd-mm-yyyy -> dd_mm_yyyy
    date_dir = os.path.join(base_output_dir, date_folder_name)
    shift_dir = os.path.join(date_dir, shift)
    os.makedirs(shift_dir, exist_ok=True)

    # file name: dd_mm_yyyy_coursecode_room_shift.xlsx
    filename = f"{date_folder_name}_{course_code}_{room_no}_{shift}.xlsx"
    filepath = os.path.join(shift_dir, filename)

    # workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{course_code} Room {room_no}"

    # merged header
    header_text = (
        f"Course: {course_code} | Room: {room_no} | Date: {date_str} | Session: {shift}"
    )
    ws.merge_cells("A1:C1")
    ws["A1"] = header_text
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # column headers
    ws["A2"] = "Roll"
    ws["B2"] = "Student Name"
    ws["C2"] = "Signature"
    ws["A2"].font = ws["B2"].font = ws["C2"].font = Font(bold=True)

    # students
    row_start = 3
    for i, (_, row) in enumerate(students_df.iterrows(), start=row_start):
        ws.cell(row=i, column=1, value=row["rollno"])
        ws.cell(row=i, column=2, value=row["Name"])
        # Signature blank

    # TA rows
    ta_start = row_start + len(students_df)
    for idx in range(5):
        ws.cell(row=ta_start + idx, column=1, value=f"TA{idx+1}")

    # Invigilator rows
    inv_start = ta_start + 5
    for idx in range(5):
        ws.cell(row=inv_start + idx, column=1, value=f"Invigilator{idx+1}")

    # widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18

    wb.save(filepath)


# ---------- export all xlsx for ONE date + shift ----------
def export_shift_excels(result, date_str, shift, room_alloc, base_output_dir="output"):
    """
    For given date+shift and room_alloc from allocate_shift_for_date(),
    create ONE xlsx per (course_code, room_no).

    File path: output/dd_mm_yyyy/Shift/dd_mm_yyyy_coursecode_room_shift.xlsx
    """
    if not room_alloc:
        return

    slot_courses = result[date_str][shift]   # course_code -> info
    any_room = next(iter(room_alloc.values()))
    day_val = any_room["day"]

    for room_no, info in room_alloc.items():
        room_df = info["df"].reset_index(drop=True)
        room_rolls = set(room_df["rollno"])

        # for every course in this shift, see if some of its students are in this room
        for course_code, cinfo in slot_courses.items():
            course_rolls = set(cinfo["df"]["rollno"])
            common_rolls = room_rolls & course_rolls
            if not common_rolls:
                continue

            # students of THIS course in THIS room
            sub_df = room_df[room_df["rollno"].isin(common_rolls)].copy()
            sub_df = sub_df.sort_values("rollno").reset_index(drop=True)

            write_room_excel(
                base_output_dir=base_output_dir,
                date_str=date_str,
                day=day_val,
                shift=shift,
                room_no=room_no,
                block=info["block"],
                course_code=course_code,
                students_df=sub_df,
            )

def generate_seats_left_excel(result,
                              rooms_df,
                              buffer,
                              mode,
                              output_dir="output",
                              filename="op_seats_left.xlsx"):
    """
    Create op_seats_left.xlsx with columns:
    Room No., Exam Capacity, Block, Allotted, Vacant (B-C)

    Allotted = max number of students ever sitting in that room
               in any single date/shift.
    Vacant   = Exam Capacity - Allotted
    """

    # clean base rooms info
    base = rooms_df.copy()
    base["Room No."] = base["Room No."].astype(str).str.strip()
    base["Block"] = base["Block"].astype(str).str.strip()
    base["Exam Capacity"] = base["Exam Capacity"].astype(int)

    # initial stats per room
    stats = {
        row["Room No."]: {
            "Exam Capacity": int(row["Exam Capacity"]),
            "Block": row["Block"],
            "Allotted": 0
        }
        for _, row in base.iterrows()
    }

    # run allocations for all dates & shifts, update max Allotted per room
    for date_str in result.keys():
        for shift in ["Morning", "Evening"]:
            room_alloc = allocate_shift_for_date(
                result, date_str, shift, rooms_df, buffer, mode
            )
            if not room_alloc:
                continue

            for room_no, info in room_alloc.items():
                count = int(info["count"])
                if room_no not in stats:
                    # in case some room appears only in allocation
                    stats[room_no] = {
                        "Exam Capacity": count,  # fallback
                        "Block": info["block"],
                        "Allotted": count,
                    }
                else:
                    stats[room_no]["Allotted"] = max(
                        stats[room_no]["Allotted"], count
                    )

    # build output rows in the same order as rooms_df
    rows = []
    for _, row in base.iterrows():
        room_no = row["Room No."]
        cap = stats[room_no]["Exam Capacity"]
        block = stats[room_no]["Block"]
        allotted = stats[room_no]["Allotted"]
        vacant = cap - allotted

        rows.append({
            "Room No.": room_no,
            "Exam Capacity": cap,
            "Block": block,
            "Allotted": allotted,
            "Vacant (B-C)": vacant,
        })

    df_out = pd.DataFrame(
        rows,
        columns=["Room No.", "Exam Capacity", "Block", "Allotted", "Vacant (B-C)"],
    )

    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, filename)
    df_out.to_excel(out_path, index=False, sheet_name="op_seats_left")
    print(f"Seats-left file written to: {out_path}")

#===============this is the end ==================
def export_shift_excels(result, date_str, shift, room_alloc,
                        base_output_dir="output",
                        photos_dir="photos",
                        logger=None):

    if not room_alloc:
        return

    if logger is None:
        logger = get_pdf_logger(base_output_dir)

    slot_courses = result[date_str][shift]
    any_room = next(iter(room_alloc.values()))
    day_val = any_room["day"]
# def export_shift_excels(result, date_str, shift, room_alloc,
#                         base_output_dir="output", photos_dir="photos"):
#     if not room_alloc:
#         return
#
#     logger = get_pdf_logger(base_output_dir)
#     slot_courses = result[date_str][shift]
#     any_room = next(iter(room_alloc.values()))
#     day_val = any_room["day"]

    for room_no, info in room_alloc.items():
        room_df = info["df"].reset_index(drop=True)
        room_rolls = set(room_df["rollno"])

        # for each course, slice students of that course in this room
        for course_code, cinfo in slot_courses.items():
            course_rolls = set(cinfo["df"]["rollno"])
            common = room_rolls & course_rolls
            if not common:
                continue

            sub_df = room_df[room_df["rollno"].isin(common)].copy()
            sub_df = sub_df.sort_values("rollno").reset_index(drop=True)

            # 1) write XLSX (as you already do)
            write_room_excel(
                base_output_dir=base_output_dir,
                date_str=date_str,
                day=day_val,
                shift=shift,
                room_no=room_no,
                block=info["block"],
                course_code=course_code,
                students_df=sub_df,
            )

            # 2) create PDF
            create_attendance_pdf_for_room(
                students_df=sub_df,
                date_str=date_str,
                day=day_val,
                shift=shift,
                room_no=str(room_no),
                course_code=course_code,
                photos_dir=photos_dir,
                output_dir=base_output_dir,
                logger=logger,
            )

def get_pdf_logger(output_dir="output"):
    os.makedirs(output_dir, exist_ok=True)
    log_path = os.path.join(output_dir, "pdf_errors.log")

    logger = logging.getLogger("pdf_generation")
    logger.setLevel(logging.INFO)

    # avoid adding multiple handlers if function called many times
    if not logger.handlers:
        fh = logging.FileHandler(log_path, mode="a", encoding="utf-8")
        fmt = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
        fh.setFormatter(fmt)
        logger.addHandler(fh)

    return logger


def create_attendance_pdf_for_room(
    students_df: pd.DataFrame,
    date_str: str,          # "dd-mm-yyyy"
    day: str,               # "Sunday"
    shift: str,             # "Morning" / "Evening"
    room_no: str,
    course_code: str,
    photos_dir: str,
    output_dir: str = "output",
    logger: logging.Logger | None = None,
):
    """
    Generate one A4 PDF like the sample for a single (date, shift, room, course).

    students_df: dataframe with columns ["rollno", "Name"] (only students of THIS course+room)
    """

    if logger is None:
        logger = get_pdf_logger(output_dir)

    try:
        # clean data
        df = students_df.copy()
        df["rollno"] = df["rollno"].astype(str).str.strip()
        df["Name"] = df["Name"].fillna("(name not found)").astype(str).str.strip()

        # PDF name: YYYY_MM_DD_<SESSION - shift>_<ROOM>_<SUBCODE>.PDF
        dt = datetime.strptime(date_str, "%d-%m-%Y")
        ymd = dt.strftime("%Y_%m_%d")
        session_part = f"SESSION-{shift}"
        pdf_name = f"{ymd}_{session_part}_{room_no}_{course_code}.PDF"

        # output path: put directly under output/date/shift or just output/ (your choice)
        # here: output/date/shift/...
        date_folder = date_str.replace("-", "_")
        pdf_dir = os.path.join(output_dir, date_folder, shift)
        os.makedirs(pdf_dir, exist_ok=True)
        pdf_path = os.path.join(pdf_dir, pdf_name)

        logger.info(f"Generating PDF: {pdf_path}")

        # --- page / layout constants ---
        page_w, page_h = A4
        margin_top = 20 * mm
        margin_bottom = 20 * mm
        header_h = 35 * mm         # area occupied by title + header lines
        table_h = 40 * mm          # invigilator table height at bottom
        row_h = 35 * mm            # height of one student row (with 3 boxes)

        # how many rows fit in different page types
        usable_only = page_h - margin_top - margin_bottom - header_h - table_h
        usable_first = page_h - margin_top - margin_bottom - header_h
        usable_mid = page_h - margin_top - margin_bottom
        usable_last = page_h - margin_top - margin_bottom - table_h

        rows_only = int(usable_only // row_h)
        rows_first = int(usable_first // row_h)
        rows_mid = int(usable_mid // row_h)
        rows_last = int(usable_last // row_h)

        cards_per_row = 3
        total_students = len(df)
        total_rows = ceil(total_students / cards_per_row)

        # decide page structure
        pages = []  # list of (rows_on_page, has_header, has_table)
        if total_rows <= max(rows_only, 1):
            # single page (header + table)
            pages.append((total_rows, True, True))
        else:
            # first page
            first_rows = max(rows_first, 1)
            pages.append((min(first_rows, total_rows), True, False))
            remaining = total_rows - first_rows

            while remaining > rows_last:
                # middle pages
                r = max(rows_mid, 1)
                take = min(r, remaining - rows_last)
                pages.append((take, False, False))
                remaining -= take

            # last page
            pages.append((remaining, False, True))

        # helper: draw header (only first page)
        # def draw_header(c: canvas.Canvas, student_count: int):
        #     c.setFont("Helvetica-Bold", 18)
        #     title = "IITP Attendance System"
        #     c.drawCentredString(page_w / 2.0, page_h - margin_top + 5 * mm, title)
        #
        #     y = page_h - margin_top - 5 * mm
        #     c.setFont("Helvetica-Bold", 10)
        #     header1 = (
        #         f"Date: {date_str} ({day}) | Shift: {shift} | "
        #         f"Room No: {room_no} | Student count: {student_count}"
        #     )
        #     c.drawString(10 * mm, y, header1)
        #
        #     y -= 6 * mm
        #     header2 = (
        #         f"Subject: {course_code} | Stud Present: "
        #         f" | Stud Absent: "
        #     )
        #     c.drawString(10 * mm, y, header2)
        def draw_header(c, student_count):
            title_y = page_h - margin_top
            c.setFont("Helvetica-Bold", 18)
            c.drawCentredString(page_w / 2.0, title_y, "IITP Attendance System")

            # a bit of extra vertical space after title
            y = title_y - 10 * mm

            c.setFont("Helvetica-Bold", 10)
            line1 = (
                f"Date: {date_str} ({day}) | Shift: {shift} | "
                f"Room No: {room_no} | Student count: {student_count}"
            )
            c.drawString(10 * mm, y, line1)

            y -= 7 * mm
            line2 = (
                f"Subject: {course_code} | Stud Present: "
                f" | Stud Absent: "
            )
            c.drawString(10 * mm, y, line2)

        # helper: draw one student box
        def draw_student_box(c, x_left, y_top, box_w, box_h, img_path, name, roll):
            # outer border
            c.rect(x_left, y_top - box_h, box_w, box_h)

            # image area
            img_margin = 2 * mm
            img_box_w = box_w * 0.32
            img_box_h = box_h - 2 * img_margin

            img_x = x_left + img_margin
            img_y = y_top - box_h + img_margin

            try:
                img = ImageReader(img_path)
                # keep aspect ratio, fit inside img_box
                iw, ih = img.getSize()
                scale = min(img_box_w / iw, img_box_h / ih)
                iw *= scale
                ih *= scale
                c.drawImage(
                    img,
                    img_x + (img_box_w - iw) / 2.0,
                    img_y + (img_box_h - ih) / 2.0,
                    iw,
                    ih,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception as e:
                # if even nopic.jpg fails, just ignore image
                logger.warning(f"Problem drawing image {img_path}: {e}")

            # text area
            text_x = x_left + img_box_w + 3 * mm
            text_y = y_top - 8 * mm

            c.setFont("Helvetica-Bold", 9)
            c.drawString(text_x, text_y, name)

            c.setFont("Helvetica-Bold", 8)
            text_y -= 6 * mm
            c.drawString(text_x, text_y, f"Roll: {roll}")

            text_y -= 6 * mm
            c.drawString(text_x, text_y, "Sign: ")
            # underline for sign
            line_y = text_y - 1 * mm
            c.line(text_x + 25, line_y, x_left + box_w - 4 * mm, line_y)

        # helper: draw invigilator table (only last page, bottom)
        # def draw_invigilator_table(c):
        #     table_top = margin_bottom + table_h
        #     # heading
        #     c.setFont("Helvetica-Bold", 10)
        #     c.drawCentredString(
        #         page_w / 2.0, table_top + 5 * mm, "Invigilator Name & Signature"
        #     )
        #
        #     rows = 10
        #     col_sl_w = 20 * mm
        #     col_name_w = 90 * mm
        #     col_sig_w = page_w - 2 * 10 * mm - col_sl_w - col_name_w
        #
        #     left_x = 10 * mm
        #     top_y = table_top - 2 * mm
        #
        #     # outer border
        #     total_h = rows * 5 * mm + 10 * mm
        #     c.rect(left_x, margin_bottom, col_sl_w + col_name_w + col_sig_w, total_h)
        #
        #     # column headers
        #     c.setFont("Helvetica-Bold", 9)
        #     y = top_y
        #     c.line(left_x, y, left_x + col_sl_w + col_name_w + col_sig_w, y)
        #     c.line(left_x + col_sl_w, y, left_x + col_sl_w, margin_bottom)
        #     c.line(left_x + col_sl_w + col_name_w, y,
        #            left_x + col_sl_w + col_name_w, margin_bottom)
        #
        #     c.drawString(left_x + 2 * mm, y - 4 * mm, "Sl No.")
        #     c.drawString(left_x + col_sl_w + 2 * mm, y - 4 * mm, "Name")
        #     c.drawString(left_x + col_sl_w + col_name_w + 2 * mm, y - 4 * mm, "Signature")
        #
        #     # row lines
        #     current_y = y - 10 * mm
        #     for _ in range(rows):
        #         c.line(left_x, current_y, left_x + col_sl_w + col_name_w + col_sig_w, current_y)
        #         current_y -= 5 * mm

        def draw_invigilator_table(c):
            rows = 10
            row_h = 8 * mm
            header_h = 8 * mm

            left_x = 20 * mm
            right_x = page_w - 20 * mm
            width = right_x - left_x

            bottom_y = margin_bottom + 5 * mm
            top_y = bottom_y + header_h + rows * row_h

            # outer border
            c.rect(left_x, bottom_y, width, top_y - bottom_y)

            # title
            c.setFont("Helvetica-Bold", 10)
            c.drawCentredString(
                (left_x + right_x) / 2.0,
                top_y + 3 * mm,
                "Invigilator Name & Signature"
            )

            # column boundaries
            col_sl_w = 20 * mm
            col_name_w = 90 * mm
            col_sig_w = width - col_sl_w - col_name_w

            x_sl = left_x
            x_name = left_x + col_sl_w
            x_sig = left_x + col_sl_w + col_name_w

            # header row line
            header_y = top_y - header_h
            c.line(left_x, header_y, right_x, header_y)

            # vertical lines
            c.line(x_name, bottom_y, x_name, top_y)
            c.line(x_sig, bottom_y, x_sig, top_y)

            # header text
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x_sl + 2 * mm, header_y + 2 * mm, "Sl No.")
            c.drawString(x_name + 2 * mm, header_y + 2 * mm, "Name")
            c.drawString(x_sig + 2 * mm, header_y + 2 * mm, "Signature")

            # horizontal lines for rows
            y = header_y - row_h
            for _ in range(rows - 1):  # we already have header separation
                c.line(left_x, y, right_x, y)
                y -= row_h

        # ------------ actual drawing loop ------------
        c = canvas.Canvas(pdf_path, pagesize=A4)
        student_index = 0

        # calculate student count (only actual students, no TA/Invigilators)
        student_count = len(df)

        for i, (rows_on_page, has_header, has_table) in enumerate(pages):
            is_first_page = (i == 0)

            if has_header:
                draw_header(c, student_count)

            # area top for rows, depending on header/table
            if is_first_page:
                y_top_rows = page_h - margin_top - header_h - 5 * mm
            else:
                y_top_rows = page_h - margin_top

            # card width
            usable_w = page_w - 2 * 10 * mm
            box_w = usable_w / cards_per_row
            x_start = 10 * mm

            y = y_top_rows

            for r in range(rows_on_page):
                # y is top of this row
                for col in range(cards_per_row):
                    if student_index >= total_students:
                        break

                    student = df.iloc[student_index]
                    roll = student["rollno"]
                    name = student["Name"] or "(name not found)"

                    # pick photo
                    # roll_img = os.path.join(photos_dir, f"{roll}.jpg")
                    # if not os.path.exists(roll_img):
                    #     roll_img = os.path.join(photos_dir, "nopic.jpg")
                    roll_img = get_photo_path(roll, photos_dir)

                    x = x_start + col * box_w
                    draw_student_box(c, x, y, box_w, row_h, roll_img, name, roll)

                    student_index += 1

                y -= row_h

            if has_table:
                draw_invigilator_table(c)

            if i < len(pages) - 1:
                c.showPage()

        c.save()
        logger.info(f"PDF created successfully: {pdf_path}")

    except Exception as e:
        logger.exception(
            f"Error generating PDF for date={date_str}, shift={shift}, "
            f"room={room_no}, course={course_code}: {e}"
        )


if __name__ == '__main__':


    # Example usage:
    excel_path = "input_data_tt.xlsx"
    result = build_course_dicts(excel_path)
    # print(list(result.keys()))

    #Pretty-print just the structure (not the full dataframes)
    # preview_structure1(result)

    # Print only counts per course (no dataframe)
    # preview_counts(result)

    # 3.Print summary for a single date (clean)
    # preview_date(result,"30-04-2016")

    #Print only course names + student counts in table-like format
    # preview_table(result)

    clash_list = detect_clashes(result)

    # Print result summary
    if not clash_list:
        print("\n✅ No clashes found.")
    else:
        print("\n❌ Clashes detected:")
        for c in clash_list:
            print("\n------ CLASH ------")
            print("Date :", c["date"])
            print("Shift:", c["shift"])
            print("Courses:", c["course1"], "<->", c["course2"])
            print("Common Rolls:", c["common_rolls"])
            print("Count:", c["count"])


    #========================================================

    date = "30-04-2016"
    shift = "Morning"
    mode = "Dense"
    # mode = "Sparse"
    buffer = 5

    preview_date(result, date)

    sorted_courses = get_sorted_courses_for_slot(result, date, shift)

    # print(type(sorted_courses))
    # print(sorted_courses)

    for c in sorted_courses:
        print(c["course_code"], c["count"])

    # ---- load and sort rooms by capacity within each building (Block) ----
    excel_path = "input_data_tt.xlsx"
    rooms_df = pd.read_excel(excel_path, sheet_name="in_room_capacity")

    # example buffer value
    rooms_sorted = prepare_rooms(rooms_df, buffer)

    # example call:
    can_allocate_shift(result, date, shift, rooms_sorted)

    # -------------------------------------------------
    # example usage
    # -------------------------------------------------
    # allocation = allocate_shift_for_date(result, date, shift, rooms_df, buffer, mode)
    # for room, info in allocation.items():
    #     print(room, info["block"], info["count"])
    #     # print(info["df"].head())
    #     print(info["df"])

    # -------------------------------------------------
    # Example usage for ALL dates and both shifts
    # -------------------------------------------------
    # rooms_df = pd.read_excel("input_data_tt.xlsx", sheet_name="in_room_capacity")
    # buffer = 5
    # mode = "Sparse"

    #================================================
    # for date_str in result.keys():
    #     for shift in ["Morning", "Evening"]:
    #         # allocate rooms for this date+shift
    #         room_alloc = allocate_shift_for_date(result, date_str, shift, rooms_df, buffer, mode)
    #         export_shift_excels(result, date_str, shift, room_alloc, base_output_dir="output")
    #

    # ---------- example loop over all dates & shifts ----------
    for date_str in result.keys():
        for shift in ["Morning", "Evening"]:
            room_alloc = allocate_shift_for_date(result, date_str, shift, rooms_df, buffer, mode)
            export_shift_excels(result, date_str, shift, room_alloc, base_output_dir="output")
    generate_overall_seating_excel(result, rooms_df, buffer, mode, output_dir="output")
    generate_seats_left_excel(result, rooms_df, buffer, mode, output_dir="output")

    logger = get_pdf_logger("output")  # add this once before loop

    for date_str in result.keys():
        for shift in ["Morning", "Evening"]:
            room_alloc = allocate_shift_for_date(result, date_str, shift, rooms_df, buffer, mode)

            export_shift_excels(
                result,
                date_str,
                shift,
                room_alloc,
                base_output_dir="output",
                photos_dir="photos",
                logger=logger
            )

    generate_overall_seating_excel(result, rooms_df, buffer, mode, output_dir="output")
    generate_seats_left_excel(result, rooms_df, buffer, mode, output_dir="output")






















